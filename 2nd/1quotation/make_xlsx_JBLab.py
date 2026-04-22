"""
견적서 XLSX 생성 - JB Lab
날짜: 2026년 4월 21일 | 품목: 접착제 A / EA / 10 / 100,000원
모든 금액 셀은 Excel 수식으로 자동 계산됨
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image as XLImage

wb = Workbook()
ws = wb.active
ws.title = "견적서"

# ── 스타일 헬퍼 ────────────────────────────────────────────────────────────────
def F(size=10, bold=False, color="000000"):
    return Font(name="맑은 고딕", size=size, bold=bold, color=color)

def A(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

T  = Side(style="thin")
M  = Side(style="medium")
NO = Side(style=None)

def B(l=None, r=None, t=None, b=None):
    return Border(left=l or NO, right=r or NO, top=t or NO, bottom=b or NO)

def cell(row, col, val=None, font=None, align=None, bdr=None, fmt=None):
    c = ws.cell(row=row, column=col)
    if val is not None:
        c.value = val
    if font:
        c.font  = font
    if align:
        c.alignment = align
    if bdr:
        c.border = bdr
    if fmt:
        c.number_format = fmt
    return c

def merge(r1, c1, r2, c2, val=None, font=None, align=None, bdr=None, fmt=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    return cell(r1, c1, val, font, align, bdr, fmt)

# ── 컬럼 폭 (48개 균일) ────────────────────────────────────────────────────────
for i in range(1, 49):
    ws.column_dimensions[get_column_letter(i)].width = 2.1

# ── 행 높이 ───────────────────────────────────────────────────────────────────
ROW_H = {1: 9.9, 2: 40.8, 3: 9.9, 4: 9.9,
         5: 24.9, 6: 24.9, 7: 24.9, 8: 24.9, 9: 24.9,
         10: 37.5, 11: 12.9, 12: 12.9,
         41: 31.5, 42: 18.0, 43: 18.0}
for r, h in ROW_H.items():
    ws.row_dimensions[r].height = h
for r in range(13, 41):
    ws.row_dimensions[r].height = 18.0

# ════════════════════════════════════════════════════════════════════
# 1. 제목 "견  적  서"
# ════════════════════════════════════════════════════════════════════
merge(1, 16, 3, 36,
      val="견  적  서",
      font=F(26, True),
      align=A("center", "center"),
      bdr=B(t=M))

merge(2, 3, 2, 4, val="NO.", font=F(10), align=A())
merge(2, 5, 2, 14, val="", font=F(10), align=A(), bdr=B(b=M))

# ════════════════════════════════════════════════════════════════════
# 2. 날짜 + 수신처 영역 (rows 5-9, cols 1-20)
# ════════════════════════════════════════════════════════════════════
# row 5: 날짜 - 좌측 상단 모서리 + 날짜 셀
cell(5, 1, bdr=B(l=M, t=M))
cell(5, 2, bdr=B(t=M))
merge(5, 3, 5, 5,   val=2026, font=F(10), align=A(), bdr=B(t=M, b=T))
cell(5, 6,  "년",   F(10), A(), B(t=M))
merge(5, 7, 5, 9,   val=4,    font=F(10), align=A(), bdr=B(t=M, b=T))
cell(5, 10, "월",   F(10), A(), B(t=M))
merge(5, 11, 5, 13, val=21,   font=F(10), align=A(), bdr=B(t=M, b=T))
cell(5, 14, "일",   F(10), A(), B(t=M))

# row 6: 좌측 테두리
cell(6, 1, bdr=B(l=M))

# row 7: 수신처
cell(7, 1, bdr=B(l=M))
merge(7, 3, 7, 16, val="JB Lab",
      font=F(12, True), align=A("center", "center"), bdr=B(b=M))
cell(7, 17, "귀  중", F(12, True), A("left", "center"))

# row 8: 좌측 테두리
cell(8, 1, bdr=B(l=M))

# row 9: 안내문 + 하단 테두리
cell(9, 1, bdr=B(l=M, b=M))
cell(9, 2, bdr=B(b=M))
merge(9, 3, 9, 20,
      val="아래와 같이 견적 드립니다.",
      font=F(10), align=A("left", "center"), bdr=B(b=M))

# ════════════════════════════════════════════════════════════════════
# 3. 공급자 정보 (rows 5-9, cols 21-48)
# ════════════════════════════════════════════════════════════════════
merge(5, 21, 9, 22,
      val="공\n\n급\n\n자",
      font=F(10), align=A("center", "center", True),
      bdr=B(l=M, t=M, b=M, r=T))

def sup_row(r, label1, val1, label2=None, val2=None):
    t = M if r == 5 else T
    b = M if r == 9 else T

    merge(r, 23, r, 29, val=label1, font=F(9), align=A(),
          bdr=B(l=T, t=t, b=b, r=T))

    if label2 is not None:
        merge(r, 30, r, 36, val=val1, font=F(9), align=A("left", "center"),
              bdr=B(l=T, t=t, b=b, r=T))
        merge(r, 37, r, 39, val=label2, font=F(9), align=A(),
              bdr=B(l=T, t=t, b=b, r=T))
        merge(r, 40, r, 48, val=val2, font=F(9), align=A("left", "center"),
              bdr=B(l=T, t=t, b=b, r=M))
    else:
        merge(r, 30, r, 48, val=val1, font=F(9), align=A("left", "center", True),
              bdr=B(l=T, t=t, b=b, r=M))

sup_row(5, "상  호  명",   "(주) 케이에스모듈테크",
           "사업자번호",    "808-81-03318")
sup_row(6, "상호(대리점)", "",
           "대    표",      "김 복 기  (인)")
sup_row(7, "주       소",
        "부산시 금정구 부산대학로 63번길 2, 부산대학교 물리학과(장전동) 307호")
sup_row(8, "담       당",  "",
           "전    화",      "")
sup_row(9, "업       태",
        "제조업, 판매업, 설계업, 통신판매업")

# ════════════════════════════════════════════════════════════════════
# 4. 합계 금액 행 (row 10) - 수식으로 자동 계산
#    AC41 = 공급가액합계(SUM 수식셀), AN41 = 세액합계(SUM 수식셀)
# ════════════════════════════════════════════════════════════════════
merge(10, 3,  10, 12, val="(공급가액+세액)",
      font=F(9), align=A(), bdr=B(l=M, t=M, b=M))
merge(10, 13, 10, 14, val="원",
      font=F(9), align=A(), bdr=B(t=M, b=M))
merge(10, 15, 10, 34, val="=AC41+AN41",
      font=F(14, True), align=A("right", "center"),
      bdr=B(t=M, b=M), fmt="#,##0")
merge(10, 35, 10, 36, val="원",
      font=F(9), align=A(), bdr=B(t=M, b=M))
cell(10, 37, "( \\",  F(9), A(), B(t=M, b=M))
cell(10, 38, bdr=B(t=M, b=M))
cell(10, 39, bdr=B(t=M, b=M))
merge(10, 40, 10, 47, val="=AC41+AN41",
      font=F(9), align=A("right", "center"),
      bdr=B(t=M, b=M), fmt="#,##0")
cell(10, 48, ")", F(9), A(), B(t=M, b=M, r=M))

# ════════════════════════════════════════════════════════════════════
# 5. 품목 테이블 헤더 (rows 11-12)
# ════════════════════════════════════════════════════════════════════
HF = F(9, True)
HA = A("center", "center")

merge(11, 1,  12, 13, "품  목  명",       HF, HA, B(l=M, t=M, b=M, r=T))
merge(11, 14, 12, 17, "단 위",             HF, HA, B(l=T, t=M, b=M, r=T))
merge(11, 18, 12, 20, "수 량",             HF, HA, B(l=T, t=M, b=M, r=T))
merge(11, 21, 12, 28, "단 가",             HF, HA, B(l=T, t=M, b=M, r=T))
merge(11, 29, 12, 39, "공  급  가  액",    HF, HA, B(l=M, t=M, b=M, r=M))
merge(11, 40, 12, 48, "세 액",             HF, HA, B(l=M, t=M, b=M, r=M))

# ════════════════════════════════════════════════════════════════════
# 6. 품목 행 (rows 13-40) - 수식으로 자동 계산
# ════════════════════════════════════════════════════════════════════
# 실제 품목 데이터
ITEMS = [
    ("접착제 A", "EA", 10, 100000),
]

IF = F(9)
for r in range(13, 41):
    idx = r - 13
    if idx < len(ITEMS):
        name, unit, qty, price = ITEMS[idx]
    else:
        name, unit, qty, price = "", "", None, None

    # 품목명: A-M (1-13)
    merge(r, 1,  r, 13, val=name, font=IF, align=A("left", "center"),
          bdr=B(l=M, t=T, b=T, r=T))
    # 단위: N-Q (14-17)
    merge(r, 14, r, 17, val=unit, font=IF, align=A("center", "center"),
          bdr=B(l=T, t=T, b=T, r=T))
    # 수량: R-T (18-20)
    merge(r, 18, r, 20, val=qty, font=IF, align=A("center", "center"),
          bdr=B(l=T, t=T, b=T, r=T), fmt="#,##0")
    # 단가: U-AB (21-28)
    merge(r, 21, r, 28, val=price, font=IF, align=A("right", "center"),
          bdr=B(l=T, t=T, b=T, r=T), fmt="#,##0")
    # 공급가액 = 수량 × 단가 (수식, R열=18, U열=21)
    # openpyxl은 열 문자로 수식을 씀: R=R열, U=U열
    supply_formula = f"=IF(R{r}*U{r}=0,\"\",R{r}*U{r})"
    merge(r, 29, r, 39, val=supply_formula, font=IF, align=A("right", "center"),
          bdr=B(l=M, t=T, b=T, r=M), fmt="#,##0")
    # 세액 = 공급가액 × 10% (수식, AC열=29)
    tax_formula = f"=IF(AC{r}=\"\",\"\",ROUND(AC{r}*0.1,0))"
    merge(r, 40, r, 48, val=tax_formula, font=IF, align=A("right", "center"),
          bdr=B(l=M, t=T, b=T, r=M), fmt="#,##0")

# ════════════════════════════════════════════════════════════════════
# 7. 합계 행 (row 41) - SUM 수식
# ════════════════════════════════════════════════════════════════════
SF = F(10, True)
merge(41, 1,  41, 13, "합       계",  SF, A(), B(l=M, t=M, b=M, r=T))
merge(41, 14, 41, 17, val=None, font=SF, bdr=B(l=T, t=M, b=M, r=T))
merge(41, 18, 41, 20, val=None, font=SF, bdr=B(l=T, t=M, b=M, r=T))
merge(41, 21, 41, 28, val=None, font=SF, bdr=B(l=T, t=M, b=M, r=T))
merge(41, 29, 41, 39, val="=SUM(AC13:AC40)",
      font=SF, align=A("right", "center"),
      bdr=B(l=M, t=M, b=M, r=M), fmt="#,##0")
merge(41, 40, 41, 48, val="=SUM(AN13:AN40)",
      font=SF, align=A("right", "center"),
      bdr=B(l=M, t=M, b=M, r=M), fmt="#,##0")

# ════════════════════════════════════════════════════════════════════
# 8. 주의사항 (rows 42-43)
# ════════════════════════════════════════════════════════════════════
merge(42, 2, 42, 47,
      val="※ 상기 금액은 부가세(VAT 10%) 포함 금액입니다. 납품 조건 및 결제 방법은 협의 후 결정합니다.",
      font=F(8), align=A("left", "center"))
merge(43, 2, 43, 47,
      val="※ 제품 사양 및 가격은 사전 예고 없이 변경될 수 있습니다. 문의: 담당자에게 연락 주시기 바랍니다.",
      font=F(8), align=A("left", "center"))

# ════════════════════════════════════════════════════════════════════
# 9. 직인 이미지 삽입 (대표자명 셀 옆)
# ════════════════════════════════════════════════════════════════════
seal_path = "E:/venture/proposal/JBLab/2nd/직인.png"
if os.path.exists(seal_path):
    try:
        img = XLImage(seal_path)
        img.width  = 72
        img.height = 72
        ws.add_image(img, "AN6")
        print("직인 삽입 완료")
    except Exception as e:
        print(f"직인 삽입 실패: {e}")
else:
    print(f"직인 파일 없음: {seal_path}")

# ════════════════════════════════════════════════════════════════════
# 10. 인쇄 설정 (A4, 1페이지)
# ════════════════════════════════════════════════════════════════════
ws.page_setup.orientation  = "portrait"
ws.page_setup.paperSize    = ws.PAPERSIZE_A4
ws.page_setup.fitToPage    = True
ws.page_setup.fitToWidth   = 1
ws.page_setup.fitToHeight  = 1
ws.page_margins = PageMargins(
    left=0.39, right=0.39, top=0.47, bottom=0.39,
    header=0.2, footer=0.2
)
ws.print_area = "A1:AV43"
ws.sheet_properties.pageSetUpPr.fitToPage = True

out = "E:/venture/proposal/JBLab/2nd/3quotation_re/견적서_JBLab.xlsx"
wb.save(out)
print(f"XLSX 저장 완료: {out}")
